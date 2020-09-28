''' Much of this comes from website-tools/scrape_conferences.py, which is more likely to be in a
    better state of update '''

__author__ = "Nessa Carson"
__copyright__ = "Copyright 2020"
__version__ = "1.0"
__email__ = "nessa.carson@syngenta.com"
__status__ = "Production"

import glob
import html
import os
import urllib.request
import urllib.error
import string

import openpyxl
from openpyxl.styles import Font, Alignment

def getevent(position):
    for n, line in enumerate(page[position:]):
        if line.strip().startswith('</tr>'):
            endposition = position + n + 1
            break
    event = [p.strip() for p in page[position:endposition]]

    return event

def captureline(line):
    result = line.partition('">')[2].partition('</td>')[0]

    return result

def decodethemes(themelist):
    replacelist = [('agro', 'agrochemistry'), ('anal', 'analytical'), ('chembio', 'chem bio'), ('comp', 'computational/data'), ('edu', 'education'), ('inorg', 'inorganic/materials'), ('medchem', 'med chem'), ('policy', 'law/policy'), ('pharm', 'pharma/regulatory')]
    otherslist = ['automation', 'careers', 'diversity', 'process', 'synthesis']

    if 'all' in themelist:
        outputlist = sorted([r[1] for r in replacelist] + otherslist)
    else:
        names = [r[0] for r in replacelist]
        newthemelist = []

        for theme in themelist:
            if theme in names:
                newthemelist.append(replacelist[names.index(theme)][1])
                if theme == 'chembio':
                    newthemelist.append('med chem')
                if theme in ['medchem', 'process']:
                    newthemelist.append('synthesis')
            else:
                newthemelist.append(theme)

        outputlist = [i for n, i in enumerate(newthemelist) if i not in newthemelist[:n]]
        outputlist.sort()

    themes = (', ').join(outputlist)

    return themes

def decoderegion(region):
    continents = [('NA', 'North America'), ('SA', 'South America'), ('Aus', 'Australasia')]
    Eurlocalities = [('WEur', 'Western Europe'), ('EEur', 'Eastern Europe')]
    USAlocalities = [('WUSA', 'Western USA'), ('EUSA', 'Eastern USA'), ('CUSA', 'Central USA')]
    UKlocalities = [('NEng', 'Northern England'), ('SEng', 'Southern England'), ('NI', 'Northern Ireland')]
    regionlist = continents + Eurlocalities + USAlocalities + UKlocalities

    if region in [r[0] for r in regionlist]:
        output = regionlist[[r[0] for r in regionlist].index(region)][1]
    else:
        output = region

    return output

def openoutput():
    os.system(f'start excel "{outputfile}"')

mydir = os.getcwd()

# Scrape from the web
url = 'http://supersciencegrl.co.uk/conferences'
try:
    with urllib.request.urlopen(url) as response:
        htmlr = response.readlines()
except urllib.error.HTTPError as error:
    print(error.code, error.read)

page = []
errors = []
for h in htmlr:
    try:
        page.append(h.decode('utf-8'))
    except UnicodeDecodeError:
        errors.append(h)

lod = []
for n, p in enumerate(page):
    if p.strip().startswith('<tr class="body') and not 'postponed' in p and not 'cancelled' in p:
        event = getevent(n)
        mydict = {}
        ColumnOneEnd = False

        for n, line in [(n, line.strip()) for n, line in enumerate(event)]:
            if n == 0:
                themelist = []
                for kw in line.split(' '):
                    if kw.startswith('c') and not kw.startswith('class'):
                        themelist.append(kw[1:])
                    elif kw.startswith('l'):
                        region = kw.translate(str.maketrans('', '', string.punctuation)).strip()[1:]
                        region = decoderegion(region)
                        mydict['region'] = region
                mydict['themes'] = decodethemes(themelist)
            elif n == 1:
                mydict['url'] = line.partition('href="')[2].partition('"')[0]
                if line.endswith('</td>'):
                    title = line.partition('rel = "noopener">')[2].partition('</td>')[0]
                    title = title.replace('<span class="new-fa"><i class="fa fa-certificate" aria-hidden="true"></i></span> ', '')
                    mydict['title'] = title
                    ColumnOneEnd = True
            elif line.startswith('<span class="tooltipconf">') or line.startswith('<!--<span class="tooltipconf'):
                pass
            elif not ColumnOneEnd and line.endswith('</td>'):
                title = line.partition('</td>')[0]
                title = title.replace('<span class="new-fa"><i class="fa fa-certificate" aria-hidden="true"></i></span> ', '').replace('</a>', '')
                mydict['title'] = title.replace('&amp;', '&')
                ColumnOneEnd = True
            elif line.startswith('<td class="column2'):
                mydict['startdate'] = captureline(line).replace('&nbsp;', ' ')
            elif line.startswith('<td class="column3'):
                mydict['enddate'] = captureline(line).replace('&nbsp;', ' ')
            elif line.startswith('<td class="column4'):
                mydict['location'] = captureline(line)
            elif line.startswith('<td class="column5'):
                price = line.partition('>')[2].partition('</td>')[0]
                price = price.replace('/', '').replace('<em>', '')
                mydict['memberprice'] = html.unescape(price).replace('–', '-')
            elif line.startswith('<td class="column6'):
                price = line.partition('>')[2].partition('</td>')[0]
                price = price.replace('/', '').replace('<em>', '')
                mydict['nonmemberprice'] = html.unescape(price).replace('–', '-')
        lod.append(mydict)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Scraped from web'
outputfile = 'RS_ConferenceList.xlsx'
ws.sheet_properties.tabColor = '00B050'

headers = ['Conference', 'Start date', 'End date', 'Location', 'Member fee', 'Nonmember fee', 'Weblink', 'Themes']
for n, h in enumerate(headers):
    ws.cell(row = 1, column = n + 1).value = h
    ws.cell(row = 1, column = n + 1).font = Font(name = 'Arial', size = 10, bold = True)
    ws.cell(row = 1, column = n + 1).alignment = Alignment(wrapText = True)

for y, event in enumerate(lod):
    ws.cell(row = y + 2, column = 1).value = event['title']
    ws.cell(row = y + 2, column = 2).value = event['startdate']
    if event['enddate'] == '&mdash;':
        ws.cell(row = y + 2, column = 3).value = '-'
    else:
        ws.cell(row = y + 2, column = 3).value = event['enddate']
    ws.cell(row = y + 2, column = 4).value = event['location']
    ws.cell(row = y + 2, column = 5).value = event['memberprice']
    ws.cell(row = y + 2, column = 6).value = event['nonmemberprice']
    ws.cell(row = y + 2, column = 7).value = event['url']
    ws.cell(row = y + 2, column = 7).hyperlink = event['url']
    ws.cell(row = y + 2, column = 7).font = Font(color = '000000FF', underline = 'single')
    ws.cell(row = y + 2, column = 8).value = event['themes']

wb.active.views.sheetView[0].tabSelected = False
wb.active = ws

wb.save(outputfile)
print(f'{outputfile} saved. ')
